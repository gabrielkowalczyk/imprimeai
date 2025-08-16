from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from .forms import RegistroForm, DocumentoForm, EnderecoForm, PagamentoForm
from .models import Cliente, Documento, Pedido


import PyPDF2
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pptx import Presentation
from decimal import Decimal

def homepage(request):
    return render(request, 'homepage.html')

def register_user(request):
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            # Cria o usuário com o telefone como username e o email
            user = User.objects.create(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'],
                password=make_password(request.POST['password'])
            )
            # Cria o objeto Cliente e o associa ao novo usuário
            Cliente.objects.create(user=user, telefone=user.username)
            
            # Autentica o usuário recém-criado
            login(request, user)
            
            return redirect('homepage')
    else:
        form = RegistroForm()
    
    context = {'form': form}
    return render(request, 'register_user.html', context)

@login_required
def fazer_pedido(request):
    if request.method == 'POST':
        form = DocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.cliente = request.user.cliente
            documento.nome_arquivo = documento.arquivo.name
            
            # LÓGICA DE CONTAGEM DE PÁGINAS AQUI
            file_extension = documento.arquivo.name.split('.')[-1].lower()
            
            if file_extension == 'pdf':
                try:
                    pdf_reader = PyPDF2.PdfReader(documento.arquivo)
                    documento.qtd_paginas = len(pdf_reader.pages)
                except Exception as e:
                    documento.qtd_paginas = 0
            elif file_extension in ['jpg', 'jpeg', 'png', 'gif']:
                documento.qtd_paginas = 1
            elif file_extension == 'docx':
                try:
                    doc = DocxDocument(documento.arquivo)
                    documento.qtd_paginas = len(doc.paragraphs)
                except Exception as e:
                    documento.qtd_paginas = 0
            elif file_extension == 'xlsx':
                try:
                    workbook = load_workbook(documento.arquivo)
                    documento.qtd_paginas = len(workbook.sheetnames)
                except Exception as e:
                    documento.qtd_paginas = 0
            elif file_extension == 'pptx':
                try:
                    prs = Presentation(documento.arquivo)
                    documento.qtd_paginas = len(prs.slides)
                except Exception as e:
                    documento.qtd_paginas = 0
            else:
                documento.qtd_paginas = 0

            documento.save()

            # LÓGICA DE PREÇO ATUALIZADA
            PRECOS = {
                'A4 Padrão (75g)': {'base': 1.00, 'cor': 1.50},
                'A4 Robusto (120g)': {'base': 1.20, 'cor': 1.80},  # Preço de exemplo
                'A4 Cartolina (180g)': {'base': 1.50, 'cor': 2.50}, # Preço de exemplo
            }

            valor_por_pagina = PRECOS.get(documento.tipo_papel, {'base': 1.00, 'cor': 1.50})['base']
            if documento.eh_colorido:
                valor_por_pagina = PRECOS.get(documento.tipo_papel, {'base': 1.00, 'cor': 1.50})['cor']
            
            # Ajuste para páginas por folha e duplex
            qtd_paginas_impressas = documento.qtd_paginas
            if documento.paginas_por_folha == '2':
                qtd_paginas_impressas = (documento.qtd_paginas + 1) // 2
            elif documento.paginas_por_folha == '4':
                qtd_paginas_impressas = (documento.qtd_paginas + 3) // 4
            
            if documento.imprimir_dois_lados:
                qtd_paginas_impressas = (qtd_paginas_impressas + 1) // 2
            
            valor_final = qtd_paginas_impressas * valor_por_pagina * documento.num_copias

            # Cria um pedido associado ao documento e ao valor
            pedido = Pedido.objects.create(
                cliente=documento.cliente,
                valor_total=valor_final,
                # Outros campos do pedido
            )
            pedido.documentos.add(documento)

            # Redireciona para a página de confirmação
            return redirect('confirmar_pedido', pedido_id=pedido.id)
    else:
        form = DocumentoForm()
    
    context = {'form': form}
    return render(request, 'fazer_pedido.html', context)

@login_required
def confirmar_pedido(request, pedido_id):
    pedido = Pedido.objects.get(id=pedido_id, cliente=request.user.cliente)
    # Redireciona para a confirmação de endereço
    return redirect('confirmar_endereco', pedido_id=pedido.id)

@login_required
def confirmar_endereco(request, pedido_id):
    pedido = Pedido.objects.get(id=pedido_id, cliente=request.user.cliente)
    cliente = pedido.cliente
    
    # Adiciona a taxa de entrega uma vez, se ainda não tiver sido adicionada
    if not pedido.taxa_entrega_adicionada:
        pedido.valor_total += Decimal('10.00')
        pedido.taxa_entrega_adicionada = True
        pedido.save()

    if request.method == 'POST':
        form = EnderecoForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            # Redireciona para a página de pagamento
            return redirect('finalizar_pagamento', pedido_id=pedido.id)
    else:
        form = EnderecoForm(instance=cliente)

    context = {'form': form, 'pedido': pedido, 'taxa_entrega': Decimal('10.00')}
    return render(request, 'confirmar_endereco.html', context)

@login_required
def finalizar_pagamento(request, pedido_id):
    pedido = Pedido.objects.get(id=pedido_id, cliente=request.user.cliente)
    
    if request.method == 'POST':
        form = PagamentoForm(request.POST, instance=pedido)
        if form.is_valid():
            pedido = form.save(commit=False)
            pedido.status_pagamento = 'Pago' # Exemplo
            pedido.save()
            return redirect('pedido_finalizado')
    else:
        form = PagamentoForm(instance=pedido)
    
    context = {'form': form, 'pedido': pedido}
    return render(request, 'finalizar_pagamento.html', context)

@login_required
def pedido_finalizado(request):
    return render(request, 'pedido_finalizado.html')
@login_required
def pedido_finalizado(request):
    return render(request, 'pedido_finalizado.html')

@login_required
def finalizar_pedido(request, pedido_id):
    if request.method == 'POST':
        pedido = Pedido.objects.get(id=pedido_id, cliente=request.user.cliente)
        # Por enquanto, mudamos o status para 'Em Impressão'
        pedido.status_pedido = 'Em Impressao'
        pedido.save()
        return redirect('pedido_finalizado')
    return redirect('homepage')

@login_required
def pedido_finalizado(request):
    return render(request, 'pedido_finalizado.html')